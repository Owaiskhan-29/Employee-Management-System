
''' I am Owais Khan.  
I am using an Excel file to store data for this assignment,
as I have tried multiple methods, but my MySQL server is still 
unable to connect with Python.'''


import openpyxl 
import os



class Employee:
    def __init__(self, emp_id, name, age, salary, position="Employee"):
        self.emp_id = emp_id
        self.name = name
        self.age = age
        self.salary = salary
        self.position = position

    def promote(self, new_position):
        self.position = new_position


class Management:
    def __init__(self, file_name="employees.xlsx"):
        self.file_name = file_name
        self.employees = []
        self.load_from_excel()  # Load employees from the Excel file when the program starts

    def load_from_excel(self):
        if os.path.exists(self.file_name):
            workbook = openpyxl.load_workbook(self.file_name)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):  
                emp_id, name, age, salary, position = row
                emp = Employee(emp_id, name, age, salary, position)
                self.employees.append(emp)
            workbook.close()

    def save_to_excel(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Employee ID", "Name", "Age", "Salary", "Position"]) 
        for emp in self.employees:
            sheet.append([emp.emp_id, emp.name, emp.age, emp.salary, emp.position])
        workbook.save(self.file_name)

    def add_employee(self, emp):
        self.employees.append(emp)
        self.save_to_excel()  # Save to Excel after adding
        print(f"Employee {emp.name} added.")

    def remove_employee(self, emp_id):
        for i in self.employees:
            if i.emp_id == emp_id:
                self.employees.remove(i)
                self.save_to_excel()  
                print(f"Employee {i.name} removed.")
                return
        print("Employee not found!")

    def promote_employee(self, emp_id, new_position):
        for emp in self.employees:
            if emp.emp_id == emp_id:
                emp.promote(new_position)
                self.save_to_excel()  
                print(f"Employee {emp.name} promoted to {new_position}.")
                return
        print("Employee not found!")

    def display_employees(self):
        employee_list = ""
        if not self.employees:
            employee_list = "No employees in the record."
        else:
            for emp in self.employees:
                employee_list += f"ID: {emp.emp_id}, Name: {emp.name}, Age: {emp.age}, Position: {emp.position}\n"
        return employee_list
